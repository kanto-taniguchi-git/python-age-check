import openpyxl as excel
import datetime
import os
from openpyxl.styles import PatternFill

# ユーザーに生年月日を入力させる
name = input("お名前を入力してください：")
birth_year = int(input("生まれた年（西暦）を入力してください："))
birth_month = int(input("生まれた月を入力してください："))
birth_day = int(input("生まれた日を入力してください："))

# 今日の日付を取得
today = datetime.date.today()

# 西暦を取得
this_year = today.year

# 誕生日が来ているか確認
birth_this_year = datetime.date(this_year, birth_month, birth_day)

if today >= birth_this_year:
    # 今年の誕生日が過ぎている場合、該当セルを緑にする。
    collor_fill = PatternFill(start_color = '00FF00', end_color = '00FF00', fill_type = 'solid')
else:
    # 今年の誕生日がまだの場合、該当セルを赤にする。
    collor_fill =PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# 新規ワークブックを作成
book = excel.Workbook()

# シートを選択
sheet = book.active

# 氏名のフォルダを作成
os.makedirs(name, exist_ok = True)

# 100歳までの年齢と対応する西暦をセルに設定する
for i in range(101):
    age = i  # 満年齢
    year = this_year - i  # 生まれた年
    
    # 現在の西暦～100年前の西暦をセルに設定
    year_cell = sheet.cell(row = (i + 1), column = 1)
    year_cell.value = str(year) + "年"
    
    # 西暦に該当する年齢をセルに設定
    age_cell = sheet.cell(row = (i + 1), column = 2)
    age_cell.value = str(i) + "歳"
    
    if year == birth_year:
        year_cell.fill = collor_fill

# 入力された氏名のフォルダにファイルを保存
file_path = f"{name}/age_check.xlsx"
book.save(file_path)